import os
import time

from dotenv import load_dotenv
from google import genai
from openpyxl import load_workbook

# Load environment variables (override=True ensures it reads the latest from .env)
load_dotenv(override=True)

API_KEY = os.getenv("GEMINI_API_KEY")

if not API_KEY:
    raise ValueError(
        "API Key not found! Make sure the .env file is present and GEMINI_API_KEY is defined."
    )

# Configuration
INPUT_FILE = "Book1.xlsx"
COLUMN_NAME = "title"
MODEL_NAME = "gemini-3-flash-preview"

client = genai.Client(api_key=API_KEY)


def normalize_query(raw_query):
    # استفاده از تگ‌های ساختاریافته و تکنیک Few-Shot برای هدایت دقیق مدل Flash
    prompt = f"""شما یک متخصص داده‌های ساختاریافته و کارشناس سئو فروشگاهی (E-commerce) هستید.
وظیفه شما استانداردسازی و نرمال‌سازی کوئری‌های جستجوی محصولات برای جلوگیری از خطای سیستم کرالر (Crawler) است.

<rules>
۱. اصلاح واحدها: واحدهای اندازه‌گیری مخفف یا اشتباه را کامل کن (مثلاً «م» یا «میل» به «میلی لیتر»، «گ» به «گرم» و عباراتی مثل «۳۰ تایی» به «بسته 30 عددی»).
۲. اصلاح فینگلیش و املا: برندها و کلمات فینگلیش را به معادل صحیح و استاندارد (فارسی یا انگلیسی رایج در سایت‌ها) تبدیل کن.
۳. ساختاردهی: در صورت امکان، نام محصول را به فرمت رایج فروشگاهی مرتب کن: [نوع محصول] [نام برند] مدل [نام مدل] [ویژگی‌های کلیدی] [حجم/وزن].
۴. حذف کلمات مزاحم: کلمات اضافی سئو، صفات تبلیغاتی و عبارات نامربوط که در کرال کردن اخلال ایجاد می‌کنند را حذف کن.
۵. خروجی نهایی: فقط و فقط نام نرمال‌شده محصول را برگردان. از چاپ کردن هرگونه تگ، توضیح اضافه، یا کاراکترهای نشانه‌گذاری خودداری کن.
</rules>

<examples>
Input: "ضدآفتاب پرایم فیوژن واتر c با اس پی اف 50 40م"
Output: فلوئید ضد آفتاب پرایم مدل فیوژن واتر حاوی ویتامین سی با SPF50 حجم 40 میلی لیتر

Input: "کرم ابرسان jute هیدرا اکتیو 70 میل"
Output: کرم آبرسان ژوت مدل Hydra Active حجم 70 میلی لیتر

Input: "ژل شستشو هیدرودرم پوست چرب acne wash 150ml"
Output: ژل شستشوی صورت هیدرودرم مناسب پوست چرب مدل Acne Wash حجم 150 میلی لیتر

Input: "پودر امینو eaa ترک نوتریشن ۳۰۰گ"
Output: پودر آمینو eaa ترک نوتریشن وزن 300 گرم

Input: "قرص ویت اسکای sky woman 30 تایی"
Output: قرص اسکای وومن ویت اسکای بسته 30 عددی

Input: "مام رول مردانه شون فاقد الومینیوم 50میل"
Output: مام رول مردانه فاقد آلومینیوم کلراید 50میل شون
</examples>

<input>
{raw_query}
</input>
"""
    try:
        response = client.models.generate_content(model=MODEL_NAME, contents=prompt)
        return response.text.strip()
    except Exception as e:
        print(f"Error processing '{raw_query}': {e}")
        return None


def main():
    if not os.path.exists(INPUT_FILE):
        print(f"Error: {INPUT_FILE} not found.")
        return

    # Load the workbook and get the active worksheet
    wb = load_workbook(INPUT_FILE)
    ws = wb.active

    # Extract headers from the first row to find column indices
    headers = [cell.value for cell in ws[1]]

    if COLUMN_NAME not in headers:
        print(f"Error: Column '{COLUMN_NAME}' not found in the Excel file.")
        return

    # openpyxl uses 1-based indexing for columns
    title_col_idx = headers.index(COLUMN_NAME) + 1

    normalized_col_name = "normalized_query"
    if normalized_col_name in headers:
        normalized_col_idx = headers.index(normalized_col_name) + 1
    else:
        # If the column doesn't exist, create it at the end
        normalized_col_idx = len(headers) + 1
        ws.cell(row=1, column=normalized_col_idx, value=normalized_col_name)
        wb.save(INPUT_FILE)  # Save the new header immediately

    total_rows = ws.max_row - 1  # Subtract 1 for the header row
    print(f"Start processing {total_rows} rows...")

    # Iterate starting from row 2 (skipping header)
    for row_num in range(2, ws.max_row + 1):
        normalized_cell = ws.cell(row=row_num, column=normalized_col_idx)
        current_normalized_val = normalized_cell.value

        # Check if the cell already contains a normalized query
        if current_normalized_val is not None:
            val_str = str(current_normalized_val).strip().lower()
            if val_str != "" and val_str != "nan":
                continue

        title_cell = ws.cell(row=row_num, column=title_col_idx)
        raw_q = title_cell.value

        # Skip if the source title is empty
        if raw_q is None or str(raw_q).strip() == "":
            continue

        raw_q_str = str(raw_q)
        print(f"Processing ({row_num - 1}/{total_rows}): {raw_q_str}")

        normalized_q = normalize_query(raw_q_str)

        if normalized_q:
            # Update the specific cell in the worksheet
            normalized_cell.value = normalized_q

            # Overwrite the input file to save progress safely
            wb.save(INPUT_FILE)
            print(f"Saved: {normalized_q}")

        # Sleep to respect rate limits
        time.sleep(4)


if __name__ == "__main__":
    main()
